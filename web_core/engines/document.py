"""
web_core.engines.document - Document parsing with strategy pattern.
Each file type gets its own FileParser subclass.
"""

from __future__ import annotations

import ast
import io
import logging
import shutil
import subprocess
from pathlib import Path
from typing import Any, Dict, List, Optional, Type

from web_core.interfaces import FileParser

logger = logging.getLogger("luqi.engines.document")

MAX_FILE_SIZE_MB = 50


class PDFParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".pdf", ".PDF"}

    def parse(self, file_path: Path) -> str:
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)[:20000]
        except ImportError:
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(str(file_path))
                return "\n".join(page.extract_text() or "" for page in reader.pages)[:20000]
            except Exception as e:
                return f"[PDF error: {e}]"


class DocxParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".docx", ".DOCX", ".doc", ".DOC"}

    def parse(self, file_path: Path) -> str:
        try:
            import docx
            doc = docx.Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs)[:20000]
        except Exception as e:
            return f"[DOCX error: {e}]"


class ExcelParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".xlsx", ".XLSX", ".xls", ".XLS"}

    def parse(self, file_path: Path) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            result = []
            for sheet in wb.worksheets[:3]:
                result.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(max_row=min(100, sheet.max_row), values_only=True):
                    result.append(", ".join(str(c) for c in row if c is not None))
            return "\n".join(result)[:20000]
        except Exception as e:
            return f"[XLSX error: {e}]"


class ImageParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".png", ".PNG", ".jpg", ".JPG", ".jpeg", ".JPEG", ".bmp", ".BMP", ".webp", ".WEBP"}

    def parse(self, file_path: Path) -> str:
        return f"[Image: {file_path.name}]"


class TextParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".txt", ".TXT", ".md", ".MD", ".json", ".JSON", ".csv", ".CSV"}

    def parse(self, file_path: Path) -> str:
        try:
            return file_path.read_text(encoding="utf-8", errors="replace")[:20000]
        except Exception as e:
            return f"[Text error: {e}]"


class PythonParser(FileParser):
    @property
    def extensions(self) -> set:
        return {".py", ".PY"}

    def parse(self, file_path: Path) -> str:
        try:
            code = file_path.read_text(encoding="utf-8")
            tree = ast.parse(code)
            functions = [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]
            classes = [n.name for n in ast.walk(tree) if isinstance(n, ast.ClassDef)]
            imports = [n for n in ast.walk(tree) if isinstance(n, (ast.Import, ast.ImportFrom))]
            return (
                f"# Python File: {file_path.name}\n\n"
                f"Classes: {classes}\n"
                f"Functions: {functions}\n"
                f"Imports: {len(imports)}\n\n"
                f"```python\n{code[:5000]}\n```"
            )
        except SyntaxError:
            return file_path.read_text(encoding="utf-8")[:10000]
        except Exception as e:
            return f"[Python error: {e}]"


class DocumentEngine:
    """Orchestrates document parsing using registered FileParser strategies."""

    _PARSERS: List[Type[FileParser]] = [
        PDFParser, DocxParser, ExcelParser, ImageParser, TextParser, PythonParser
    ]

    def __init__(self, sandbox_dir: Path):
        self.sandbox_dir = sandbox_dir
        self.sandbox_dir.mkdir(parents=True, exist_ok=True)
        self._parser_map: Dict[str, FileParser] = {}
        for parser_cls in self._PARSERS:
            p = parser_cls()
            for ext in p.extensions:
                self._parser_map[ext.lower()] = p

    def parse(self, file_path: str | Path) -> Dict[str, Any]:
        src = Path(file_path)
        if not src.exists():
            return {"status": "error", "error": f"File not found: {src}"}

        size_mb = src.stat().st_size / (1024 * 1024)
        if size_mb > MAX_FILE_SIZE_MB:
            return {"status": "error", "error": f"File too large: {size_mb:.1f}MB (max {MAX_FILE_SIZE_MB}MB)"}

        ext = src.suffix.lower()
        parser = self._parser_map.get(ext)
        if not parser:
            return {"status": "error", "error": f"Unsupported format: {ext}"}

        # Copy to sandbox (skip if already there)
        secure_path = self.sandbox_dir / src.name
        if src.resolve() != secure_path.resolve():
            try:
                shutil.copy2(str(src), str(secure_path))
            except shutil.SameFileError:
                pass  # Already the same file
            except Exception as e:
                return {"status": "error", "error": f"Sandbox copy failed: {e}"}

        try:
            content = parser.parse(secure_path)
            return {"status": "ok", "filename": src.name, "type": ext, "content": content, "size_mb": round(size_mb, 2)}
        except Exception as e:
            logger.error("Parse error for %s: %s", src.name, e)
            return {"status": "error", "error": f"Parse failed: {e}"}

    def supported_extensions(self) -> set:
        return set(self._parser_map.keys())

    def get_parser_for(self, ext: str) -> Optional[FileParser]:
        return self._parser_map.get(ext.lower())
